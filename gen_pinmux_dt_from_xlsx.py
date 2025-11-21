#!/usr/bin/env python3
"""
Generate DTS pinmux from Jetson Thor pinmux template (.xlsm/.xlsx).

Integrates logic inspired by generate_pinmux.py:

1) String-based mappings for:
   - PUPD        -> TEGRA_PIN_PULL_*
   - Tristate    -> TEGRA_PIN_ENABLE / TEGRA_PIN_DISABLE
   - E_Input     -> TEGRA_PIN_ENABLE / TEGRA_PIN_DISABLE
   - DRV_TYPE    -> TEGRA_PIN_1X_DRIVER / TEGRA_PIN_2X_DRIVER /
                    TEGRA_PIN_DEFAULT_DRIVE_1X / TEGRA_PIN_DEFAULT_DRIVE_2X

2) Open-drain detection:
   - If Pin Direction == "Open-Drain" OR E_IO_OD column is "ENABLE"
     -> nvidia,open-drain = <TEGRA_PIN_ENABLE>;

3) pinmux@ac281000 structure:
   - / {
         pinmux@ac281000 {
             pinctrl-names = "default", "drive", "unused";
             pinctrl-0 = <&pinmux_default>;
             pinctrl-1 = <&drive_default>;
             pinctrl-2 = <&pinmux_unused_lowpower>;

             pinmux_default: common {
                 ...
             };

             drive_default: drive { };
             pinmux_unused_lowpower: unused_lowpower { };
         };
       };

Sheet assumptions:
- Sheet name: "Jetson Thor_DevKit" (configurable)
- Header row contains: "Pin #", "Signal Name", "MPIO"
- Row above header contains semantic titles like:
    "Device Tree Pin Name", "PUPD", "Tristate", "E_Input",
    "Customer Usage", "Pin Direction", "Req. Initial State",
    "E_IO_OD...", "DRV_TYPE...", "E_LPBK...", "Lock", etc.

Dependencies:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook


# Normalization & mapping helpers
def _norm_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_upper(v) -> str:
    return _norm_str(v).upper()


def map_pull(pupd: str) -> str:
    """
    Map PUPD column to TEGRA_PIN_PULL_* macros.

    From generate_pinmux.py:
      NORMAL    -> TEGRA_PIN_PULL_NONE
      PULL_DOWN -> TEGRA_PIN_PULL_DOWN
      PULL_UP   -> TEGRA_PIN_PULL_UP
    """
    v = _norm_upper(pupd)
    if v == "PULL_UP":
        return "TEGRA_PIN_PULL_UP"
    if v == "PULL_DOWN":
        return "TEGRA_PIN_PULL_DOWN"
    # "NORMAL", blank, etc.
    return "TEGRA_PIN_PULL_NONE"


def map_tristate(tristate: str) -> str:
    """
    Thor sheet uses 'TRISTATE' / 'NORMAL' semantics.

    generate_pinmux.py + VBA:
      - TRISTATE string => bit=1 => TEGRA_PIN_ENABLE
      - anything else   => bit=0 => TEGRA_PIN_DISABLE
    """
    v = _norm_upper(tristate)
    if v == "TRISTATE":
        return "TEGRA_PIN_ENABLE"
    return "TEGRA_PIN_DISABLE"


def map_e_input(einput: str) -> str:
    """
    E_Input column mapping.

    generate_pinmux.py:
      ENABLE  -> TEGRA_PIN_ENABLE
      others  -> TEGRA_PIN_DISABLE
    """
    v = _norm_upper(einput)
    if v == "ENABLE":
        return "TEGRA_PIN_ENABLE"
    return "TEGRA_PIN_DISABLE"


def map_enable_disable(flag: str) -> str:
    """
    Generic Enable/Disable => TEGRA_PIN_ENABLE / TEGRA_PIN_DISABLE,
    used for E_IO_OD, E_LPBK, Lock, etc.
    """
    v = _norm_upper(flag)
    if v == "ENABLE":
        return "TEGRA_PIN_ENABLE"
    return "TEGRA_PIN_DISABLE"


def map_drv_type(drv: str) -> str:
    """
    DRV_TYPE mapping based on generate_pinmux.py's get_drive_str():

      ENABLE  -> TEGRA_PIN_2X_DRIVER
      DEF_1X  -> TEGRA_PIN_DEFAULT_DRIVE_1X
      DEF_2X  -> TEGRA_PIN_DEFAULT_DRIVE_2X
      other   -> TEGRA_PIN_1X_DRIVER
    """
    v = _norm_upper(drv)
    if v == "ENABLE":
        return "TEGRA_PIN_2X_DRIVER"
    if v == "DEF_1X":
        return "TEGRA_PIN_DEFAULT_DRIVE_1X"
    if v == "DEF_2X":
        return "TEGRA_PIN_DEFAULT_DRIVE_2X"
    return "TEGRA_PIN_1X_DRIVER"


def map_open_drain(pin_dir: str, od_flag: str) -> bool:
    """
    Open-drain detection inspired by generate_pinmux.py get_od_str:

      - If Pin Direction is "OPEN-DRAIN"           -> Enable
      - OR if OD/E_IO_OD column is "ENABLE"       -> Enable
      - Else                                      -> Disable
    """
    pd = _norm_upper(pin_dir)
    od = _norm_upper(od_flag)
    if pd == "OPEN-DRAIN":
        return True
    if od == "ENABLE":
        return True
    return False


def _safe_function_name(usage: str, pin_group: str, usage_desc: str) -> str:
    """
    Decide between RSVDx and the "safe" function name, mimicking the
    Excel/VBA behavior as closely as we can from the sheet:

    - If Customer Usage starts with 'unused_'
    - AND the description says 'UNUSED'
    - AND Pin Group is RSVD0/RSVD1/RSVD2/RSVD3

      => use rsvd0/rsvd1/rsvd2/rsvd3 (from Pin Group).

    - Otherwise, use the Customer Usage string (lowercased), which is
      how the Excel templates encode e.g. SHUTDOWN_N, etc.
    """
    u = _norm_str(usage)
    u_l = u.lower()
    pg = _norm_upper(pin_group)
    desc = _norm_upper(usage_desc)

    if u_l.startswith("unused_"):
        if desc == "UNUSED" and pg.startswith("RSVD"):
            # e.g. Pin Group "RSVD1" -> function "rsvd1"
            return pg.lower()

    # Default: keep the usage as the function, lowercased
    return u_l


# Header row discovery & column mapping
def find_header_rows(ws) -> Tuple[int, int]:
    """
    Find:
      - hdr_row: the row containing 'Pin #', 'Signal Name', 'MPIO'
      - labels_row: the row just above hdr_row, containing semantic titles
    """
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        a = _norm_str(ws.cell(row=r, column=1).value)
        b = _norm_str(ws.cell(row=r, column=2).value)
        c = _norm_str(ws.cell(row=r, column=3).value)
        if a == "Pin #" and b == "Signal Name" and c == "MPIO":
            hdr_row = r
            labels_row = r - 1
            if labels_row < 1:
                raise RuntimeError(
                    "Found 'Pin # / Signal Name / MPIO' at row 1; "
                    "cannot locate labels row above."
                )
            return hdr_row, labels_row

    raise RuntimeError("Could not find 'Pin # / Signal Name / MPIO' header row.")


def build_column_map(ws, labels_row: int) -> Dict[str, int]:
    """
    Build a semantic column map from the label row.

    Keys we care about:
        pin_num      -> "Pin #"
        signal_name  -> "Signal Name"
        mpio         -> "MPIO"
        dt_name      -> "Device Tree Pin Name"
        pupd         -> "PUPD"
        tristate     -> "Tristate"
        einput       -> "E_Input"
        drv_type     -> "DRV_TYPE..."
        cust_usage   -> "Customer Usage"
        pin_dir      -> "Pin Direction"
        init_state   -> "Req. Initial State"
        eio_od       -> "E_IO_OD..."
        elpbk        -> "E_LPBK..."
        lock         -> "Lock"
        pin_group    -> "Pin Group" (for RSVD0/1/2/3 detection)
        usage_desc   -> "Customer Usage Description or Net Names"
    """
    col_map: Dict[str, int] = {}

    max_col = ws.max_column
    for c in range(1, max_col + 1):
        title = ws.cell(row=labels_row, column=c).value
        t = _norm_str(title)

        if t == "Device Tree Pin Name":
            col_map["dt_name"] = c
        elif t == "Customer Usage":
            col_map["cust_usage"] = c
        elif t == "Pin Direction":
            col_map["pin_dir"] = c
        elif t == "Req. Initial State":
            col_map["init_state"] = c
        elif t == "Lock":
            col_map["lock"] = c
        elif t.startswith("E_IO_OD"):
            col_map["eio_od"] = c
        elif t.startswith("DRV_TYPE"):
            col_map["drv_type"] = c
        elif t.startswith("E_LPBK"):
            col_map["elpbk"] = c
        elif t == "PUPD":
            col_map["pupd"] = c
        elif t == "Tristate":
            col_map["tristate"] = c
        elif t == "E_Input":
            col_map["einput"] = c
        elif t == "Pin Group":
            col_map["pin_group"] = c
        elif t == "Customer Usage Description or Net Names":
            col_map["usage_desc"] = c

    # Pin # / Signal Name / MPIO come from the header row itself
    col_map.setdefault("pin_num", 1)
    col_map.setdefault("signal_name", 2)
    col_map.setdefault("mpio", 3)

    required = [
        "dt_name",
        "pupd",
        "tristate",
        "einput",
        "cust_usage",
        "pin_dir",
        "init_state",
        "lock",
        "eio_od",
        "drv_type",
        "elpbk",
    ]
    missing = [k for k in required if k not in col_map]
    if missing:
        raise RuntimeError(
            f"Missing required columns in labels row {labels_row}: {missing}"
        )

    return col_map


# DTS generation
def generate_pinmux(ws, hdr_row: int, labels_row: int) -> str:
    """
    Generate full DTS text containing:

        / {
            pinmux@ac281000 {
                pinctrl-names = "default", "drive", "unused";
                pinctrl-0 = <&pinmux_default>;
                pinctrl-1 = <&drive_default>;
                pinctrl-2 = <&pinmux_unused_lowpower>;

                pinmux_default: common {
                    ... all pins ...
                };

                drive_default: drive { };
                pinmux_unused_lowpower: unused_lowpower { };
            };
        };

    Returns: (dts_text, emitted_count)
    """
    col = build_column_map(ws, labels_row)

    first_data_row = hdr_row + 1
    last_row = ws.max_row

    lines = []

    # Excel-like includes at top
    lines.append('/* Auto-generated from Jetson Thor pinmux spreadsheet */')
    lines.append('#include "t264-pinctrl-tegra.h"')
    # Many reference DTS also include a gpio dtsi here; keep this generic.
    lines.append('#include "tegra264-gpio.h"')
    lines.append("")
    lines.append("/ {")
    lines.append("\tpinmux@ac281000 {")
    lines.append('\t\tpinctrl-names = "default", "drive", "unused";')
    lines.append("\t\tpinctrl-0 = <&pinmux_default>;")
    lines.append("\t\tpinctrl-1 = <&drive_default>;")
    lines.append("\t\tpinctrl-2 = <&pinmux_unused_lowpower>;")
    lines.append("")
    lines.append("\t\tpinmux_default: common {")
    lines.append("\t\t\t/* SFIO + GPIO Pin Configuration (auto-generated) */")

    emitted = 0

    for r in range(first_data_row, last_row + 1):
        dt_name = _norm_str(ws.cell(row=r, column=col["dt_name"]).value)
        if not dt_name:
            continue

        # Skip power-rail rows (e.g. VDDIO_SYS)
        if dt_name.upper().startswith("VDDIO_"):
            continue

        pin_num = _norm_str(ws.cell(row=r, column=col["pin_num"]).value)
        signal_name = _norm_str(ws.cell(row=r, column=col["signal_name"]).value)
        mpio = _norm_str(ws.cell(row=r, column=col["mpio"]).value)
        usage = _norm_str(ws.cell(row=r, column=col["cust_usage"]).value)
        pin_dir = _norm_str(ws.cell(row=r, column=col["pin_dir"]).value)
        pupd = _norm_str(ws.cell(row=r, column=col["pupd"]).value)
        tristate = _norm_str(ws.cell(row=r, column=col["tristate"]).value)
        einput = _norm_str(ws.cell(row=r, column=col["einput"]).value)
        drv = _norm_str(ws.cell(row=r, column=col["drv_type"]).value)
        lock = _norm_str(ws.cell(row=r, column=col["lock"]).value)
        eio_od = _norm_str(ws.cell(row=r, column=col["eio_od"]).value)
        elpbk = _norm_str(ws.cell(row=r, column=col["elpbk"]).value)

        # Optional columns for RSVD logic
        pin_group = ""
        if "pin_group" in col:
            pin_group = _norm_str(ws.cell(row=r, column=col["pin_group"]).value)
        usage_desc = ""
        if "usage_desc" in col:
            usage_desc = _norm_str(ws.cell(row=r, column=col["usage_desc"]).value)

        # Map to macros using string-based helpers
        pull_macro = map_pull(pupd)
        tristate_macro = map_tristate(tristate)
        einput_macro = map_e_input(einput)
        drv_macro = map_drv_type(drv)
        lock_macro = map_enable_disable(lock)
        eio_macro = map_enable_disable(eio_od)
        elpbk_macro = map_enable_disable(elpbk)

        # Open-drain detection from both Pin Direction + E_IO_OD
        od_enabled = map_open_drain(pin_dir, eio_od)

        # Node name / pins string: lowercased DT pin name
        node_name = dt_name.lower()
        pins_str = dt_name.lower()

        # Function: use SafeFunctionName-style logic
        func_name = _safe_function_name(usage, pin_group, usage_desc)

        # Comment "Pin H53 - MCLK03"
        if pin_num or signal_name:
            comment = f"/* Pin {pin_num or '?'} - {signal_name or '?'} ({mpio or ''}) */"
            lines.append(f"\t\t\t{comment}")

        lines.append(f"\t\t\t{node_name} {{")
        lines.append(f'\t\t\t\tnvidia,pins = "{pins_str}";')

        if func_name:
            lines.append(f'\t\t\t\tnvidia,function = "{func_name}";')

        lines.append(f"\t\t\t\tnvidia,pull = <{pull_macro}>;")
        lines.append(f"\t\t\t\tnvidia,tristate = <{tristate_macro}>;")
        lines.append(f"\t\t\t\tnvidia,enable-input = <{einput_macro}>;")
        lines.append(f"\t\t\t\tnvidia,drv-type = <{drv_macro}>;")

        # Excel-style: always emit e-io-od and e-lpbk
        lines.append(f"\t\t\t\tnvidia,e-io-od = <{eio_macro}>;")
        lines.append(f"\t\t\t\tnvidia,e-lpbk = <{elpbk_macro}>;")

        # Lock only when ENABLE
        if lock_macro == "TEGRA_PIN_ENABLE":
            lines.append("\t\t\t\tnvidia,lock = <TEGRA_PIN_ENABLE>;")

        # Open-drain property
        if od_enabled:
            lines.append("\t\t\t\tnvidia,open-drain = <TEGRA_PIN_ENABLE>;")

        lines.append("\t\t\t};")
        emitted += 1

    # Close pinmux_default
    lines.append("\t\t};")

    # Keep placeholders for drive_default and pinmux_unused_lowpower
    lines.append("")
    lines.append("\t\tdrive_default: drive {")
    lines.append("\t\t};")
    lines.append("")
    lines.append("\t\tpinmux_unused_lowpower: unused_lowpower {")
    lines.append("\t\t};")
    lines.append("\t};")   # end pinmux@ac281000
    lines.append("};")     # end / { }

    return "\n".join(lines), emitted


# CLI
def main():
    ap = argparse.ArgumentParser(
        description="Generate Jetson Thor pinmux DTS from .xlsm/.xlsx (Excel-style)."
    )
    ap.add_argument("excel", help="Input .xlsm/.xlsx file (Thor pinmux template)")
    ap.add_argument(
        "-s",
        "--sheet",
        default="Jetson Thor_DevKit",
        help="Worksheet name (default: 'Jetson Thor_DevKit')",
    )
    ap.add_argument(
        "-o",
        "--out",
        default="pinmux-thor.dtsi",
        help="Output DTS filename (default: pinmux-thor.dtsi)",
    )
    args = ap.parse_args()

    wb = load_workbook(args.excel, data_only=True)
    if args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    hdr_row, labels_row = find_header_rows(ws)
    dts_text, count = generate_pinmux(ws, hdr_row, labels_row)

    out_path = Path(args.out)
    out_path.write_text(dts_text, encoding="utf-8")
    print(f"Wrote {out_path} with {count} pin blocks.")


if __name__ == "__main__":
    main()
